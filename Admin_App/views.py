from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from U_Auth.models import User
from django.db import IntegrityError
from django.contrib import messages
from Admin_App.models import *
from django.http import JsonResponse
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
# Create your views here.


# ---------------------------------------------- Collector -------------------------------------------------------- #

@login_required
def collector_list(request):
    collector = User.objects.filter(is_collector=True)
    context = {
        'collector' : collector,
    }
    return render(request, 'Admin/collector_list.html', context)


@login_required
def collector_add(request):
    if request.method == 'POST':
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        mobile_no = request.POST.get('mobile_no')
        is_collector = request.POST.get('is_collector')
        username = request.POST.get('username')
        password = request.POST.get('password')
        try:
            user = User.objects.create_user(
                first_name=first_name,
                last_name=last_name,
                email=email,
                mobile_no=mobile_no,
                username=username,
                password=password,
                is_collector=True,
            )
            messages.success(request, f"The username '{username}' added successfully.")
            return redirect('collector_list')
        except IntegrityError:
            messages.error(request, f"The username '{username}' is already taken. Please choose a different username.")
            return redirect('collector_add')         
    return render(request, 'Admin/collector_add.html')

@login_required
def collector_edit(request, collector_id):
    collector = User.objects.get(id=collector_id, is_collector=True)
    if request.method == 'POST':
        collector.first_name = request.POST.get('first_name')
        collector.last_name = request.POST.get('last_name')
        collector.email = request.POST.get('email')
        collector.mobile_no = request.POST.get('mobile_no')
        collector.username = request.POST.get('username')
        password = request.POST.get('password')
        if password:
            collector.set_password(password)
        try:
            collector.save()
            messages.success(request, f"Collector '{collector.username}' updated successfully.")
            return redirect('collector_list')
        except IntegrityError:
            messages.error(request, f"The username '{collector.username}' is already taken. Please choose a different username.")
            return redirect('collector_edit', collector_id=collector_id)           
    return render(request, 'Admin/collector_edit.html', {'collector': collector})


# def collector_view(request, collector_id):
#     collector = User.objects.get(id=collector_id, is_collector=True)
#     orders = Order.objects.filter(collector_name=collector).order_by('-id')
#     total_total = sum(order.total for order in orders)
#     total_money_collected = sum(order.money_collected for order in orders)
#     total_money_pending = sum(order.money_pending for order in orders)

#     context = {
#         'collector': collector,
#         'orders': orders,
        
#         'total_total': total_total,
#         'total_money_collected': total_money_collected,
#         'total_money_pending': total_money_pending,
#     }
#     return render(request, 'Admin/collector_view.html', context)

def collector_view(request, collector_id):
    collector = User.objects.get(id=collector_id, is_collector=True)

    # Get the start and end time from the request parameters
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    # Convert the start and end time strings to datetime objects
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None

    # Filter the orders based on the collector and date range
    orders = Order.objects.filter(collector_name=collector)
    if start_date:
        orders = orders.filter(date__gte=start_date)
    if end_date:
        orders = orders.filter(date__lte=end_date)

    orders = orders.order_by('-id')

    total_total = sum(order.total if order.total is not None else 0 for order in orders)
    total_money_collected = sum(order.money_collected if order.money_collected is not None else 0 for order in orders)
    total_money_pending = sum(order.money_pending if order.money_pending is not None else order.total for order in orders)

    context = {
        'collector': collector,
        'orders': orders,
        'total_total': total_total,
        'total_money_collected': total_money_collected,
        'total_money_pending': total_money_pending,
        'start_date': start_date_str,
        'end_date': end_date_str,
    }
    return render(request, 'Admin/collector_view.html', context)



@login_required
def collector_delete(request, collector_id):
    collector = User.objects.get(id=collector_id, is_collector=True)
    try:
        collector.delete()
        messages.success(request, f"Collector '{collector.username}' Delete successfully.")
        return redirect('collector_list')
    except IntegrityError:
            messages.error(request, f"The username '{collector.username}' Delete Failed.")
            return redirect('collector_list')

            
    

# ---------------------------------------------- Delivry Boy -------------------------------------------------------- #

@login_required
def delivery_boy_list(request):
    delivery_boy = User.objects.filter(is_delivery_boy=True)
    context = {
        'delivery_boy' : delivery_boy,
    }
    return render(request, 'Admin/delivery_boy_list.html', context)

@login_required
def delivery_boy_add(request):
    if request.method == 'POST':
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        mobile_no = request.POST.get('mobile_no')
        is_delivery_boy = request.POST.get('is_delivery_boy')
        username = request.POST.get('username')
        password = request.POST.get('password')
        try:
            user = User.objects.create_user(
                first_name=first_name,
                last_name=last_name,
                email=email,
                mobile_no=mobile_no,
                username=username,
                password=password,
                is_delivery_boy=True,
            )
            messages.success(request, f"The username '{username}' added successfully.")
            return redirect('delivery_boy_list')
        except IntegrityError:
            messages.error(request, f"The username '{username}' is already taken. Please choose a different username.")
            return redirect('delivery_boy_add')
    return render(request, 'Admin/delivery_boy_add.html')

@login_required
def delivery_boy_edit(request, delivery_boy_id):
    delivery_boy = User.objects.get(id=delivery_boy_id, is_delivery_boy=True)
    if request.method == 'POST':
        delivery_boy.first_name = request.POST.get('first_name')
        delivery_boy.last_name = request.POST.get('last_name')
        delivery_boy.email = request.POST.get('email')
        delivery_boy.mobile_no = request.POST.get('mobile_no')
        delivery_boy.username = request.POST.get('username')
        password = request.POST.get('password')
        if password:
            delivery_boy.set_password(password)
        try:
            delivery_boy.save()
            messages.success(request, f"Delivery Boy '{delivery_boy.username}' updated successfully.")
            return redirect('delivery_boy_list')
        except IntegrityError:
            messages.error(request, f"The username '{delivery_boy.username}' is already taken. Please choose a different username.")
            return redirect('delivery_boy_edit', delivery_boy_id=delivery_boy_id)            
    return render(request, 'Admin/delivery_boy_edit.html', {'delivery_boy': delivery_boy})


# def delivery_boy_view(request, delivery_boy_id):
#     delivery_boy = User.objects.get(id=delivery_boy_id, is_delivery_boy=True)
#     orders = Order.objects.filter(delivery_boy_name=delivery_boy).order_by('-id')
#     context = {
#         'delivery_boy': delivery_boy,
#         'orders': orders,
#     }
#     return render(request, 'Admin/delivery_boy_view.html', context)


def delivery_boy_view(request, delivery_boy_id):
    delivery_boy = User.objects.get(id=delivery_boy_id, is_delivery_boy=True)

    # Get the start and end time from the request parameters
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    # Convert the start and end time strings to datetime objects
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None

    # Filter the orders based on the delivery_boy and date range
    orders = Order.objects.filter(delivery_boy_name=delivery_boy)
    if start_date:
        orders = orders.filter(date__gte=start_date)
    if end_date:
        orders = orders.filter(date__lte=end_date)

    orders = orders.order_by('-id')


    context = {
        'delivery_boy': delivery_boy,
        'orders': orders,
        'start_date': start_date_str,
        'end_date': end_date_str,
    }
    return render(request, 'Admin/delivery_boy_view.html', context)


@login_required
def delivery_boy_delete(request, delivery_boy_id):
    delivery_boy = User.objects.get(id=delivery_boy_id, is_delivery_boy=True)
    try:
        delivery_boy.delete()
        messages.success(request, f"delivery_boy '{delivery_boy.username}' Delete successfully.")
        return redirect('delivery_boy_list')
    except IntegrityError:
            messages.error(request, f"The username '{delivery_boy.username}' Delete Failed.")
            return redirect('delivery_boy_list')

        
# ---------------------------------------------- Customer -------------------------------------------------------- #



@login_required
def customer_list(request):
    customer = Customer.objects.all()
    context = {
        'customer' : customer,
    }
    return render(request, 'Admin/customer_list.html', context)


@login_required
def customer_add(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        mobile = request.POST.get('mobile')
        address = request.POST.get('address')
        try:
            customer = Customer(
                name=name,
                mobile=mobile,
                address=address,
            )
            customer.save()
            messages.success(request, f"The '{customer.name}' add success fully")
            return redirect('customer_list')
        except IntegrityError:
            messages.error(request, f"Customeradd field")
            return redirect('customer_add')
    return render(request, 'Admin/customer_add.html')


@login_required
def customer_edit(request, customer_id):
    try:
        customer = Customer.objects.get(id=customer_id)
    except Customer.DoesNotExist:
        messages.error(request, f"customer with ID {customer_id} does not exist.")
        return redirect('customer_list')
    if request.method == 'POST':
        customer.name = request.POST.get('name')
        customer.mobile = request.POST.get('mobile')
        customer.address = request.POST.get('address')
        try:
            customer.save()
            messages.success(request, f"Delivery Boy '{customer.name}' updated successfully.")
            return redirect('customer_list')
        except IntegrityError:
            messages.error(request, f"The username '{customer.name}' is already taken. Please choose a different username.")
            return redirect('customer_edit', customer_id=customer_id)            
    return render(request, 'Admin/customer_edit.html', {'customer': customer})



def customer_view(request, customer_id):
    customer = Customer.objects.get(id=customer_id)

    # Get the start and end time from the request parameters
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    # Convert the start and end time strings to datetime objects
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None

    # Filter the orders based on customer and date range
    orders = Order.objects.filter(customer_name=customer)
    if start_date:
        orders = orders.filter(date__gte=start_date)
    if end_date:
        orders = orders.filter(date__lte=end_date)

    orders = orders.order_by('-id')

    

    total_debit_total_sum = sum(order.total for order in orders if order.money_collected == 0)
    total_credit_money_collected = sum(order.money_collected if order.money_collected is not None else 0 for order in orders)
    total_money_pending = sum(order.money_pending if order.money_pending is not None else order.total for order in orders)

    
    if 'export' in request.GET:
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=orders.xlsx'

                workbook = Workbook()
                worksheet = workbook.active

                # Write headers
                headers = ['Customer', 'Date', 'Description', 'Debit', 'Credit']
                for col_num, header in enumerate(headers, 1):
                    worksheet.cell(row=1, column=col_num, value=header)

                for col_num in range(1, len(headers) + 1):
                    column_letter = get_column_letter(col_num)
                    worksheet.column_dimensions[column_letter].width = 15

                # Write data
                for row_num, order in enumerate(orders, 2):
                    worksheet.cell(row=row_num, column=1, value=order.customer_name.name)  # Assuming name is the field representing customer name
                    worksheet.cell(row=row_num, column=2, value=order.date.strftime('%Y-%m-%d'))  # Convert date to formatted string

                    if order.money_collected == 0:
                        cell_value = "Sales"
                    else:
                        cell_value = "Cash"
                    worksheet.cell(row=row_num, column=3, value=cell_value)

                    if order.money_collected == 0:
                        cell_value_debit = order.total
                        cell_value_credit = ''
                    else:
                        cell_value_debit = ''
                        cell_value_credit = order.money_collected

                    worksheet.cell(row=row_num, column=4, value=cell_value_debit)
                    worksheet.cell(row=row_num, column=5, value=cell_value_credit)
                    


                workbook.save(response)
                return response

    context = {
        'customer': customer,
        'orders': orders,
        'total_credit_money_collected': total_credit_money_collected,
        'total_debit_total_sum' : total_debit_total_sum,
        'total_money_pending' : total_money_pending,
        'start_date': start_date_str,
        'end_date': end_date_str,
    }
    return render(request, 'Admin/customer_view.html', context)


@login_required
def customer_delete(request, customer_id):
    try:
        customer = Customer.objects.get(id=customer_id)
    except Customer.DoesNotExist:
        messages.error(request, f"customer with ID {customer_id} does not exist.")
        return redirect('customer_list')
    
    try:
        customer.delete()
        messages.success(request, f"customer '{customer.name}' Delete successfully.")
        return redirect('customer_list')
    except IntegrityError:
            messages.error(request, f"The username '{customer.name}' Delete Failed.")
            return redirect('customer_list')
    

# ---------------------------------------------- Wholesaler -------------------------------------------------------- #

@login_required
def wholesaler_list(request):
    wholesaler = Wholesaler.objects.all()
    context = {
        'wholesaler' : wholesaler,
    }
    return render(request, 'Admin/wholesaler_list.html', context)


@login_required
def wholesaler_add(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        mobile = request.POST.get('mobile')
        address = request.POST.get('address')
        try:
            wholesaler = Wholesaler(
                name=name,
                mobile=mobile,
                address=address,
            )
            wholesaler.save()
            messages.success(request, f"The '{wholesaler.name}' add success fully")
            return redirect('wholesaler_list')
        except IntegrityError:
            messages.error(request, f"wholesaleradd field")
            return redirect('wholesaler_add')
    return render(request, 'Admin/wholesaler_add.html')


@login_required
def wholesaler_edit(request, wholesaler_id):
    try:
        wholesaler = Wholesaler.objects.get(id=wholesaler_id)
    except Wholesaler.DoesNotExist:
        messages.error(request, f"wholesaler with ID {wholesaler_id} does not exist.")
        return redirect('wholesaler_list')
    if request.method == 'POST':
        wholesaler.name = request.POST.get('name')
        wholesaler.mobile = request.POST.get('mobile')
        wholesaler.address = request.POST.get('address')
        try:
            wholesaler.save()
            messages.success(request, f"Delivery Boy '{wholesaler.name}' updated successfully.")
            return redirect('wholesaler_list')
        except IntegrityError:
            messages.error(request, f"The username '{wholesaler.name}' is already taken. Please choose a different username.")
            return redirect('wholesaler_edit', wholesaler_id=wholesaler_id)            
    return render(request, 'Admin/wholesaler_edit.html', {'wholesaler': wholesaler})

# def wholesaler_view(request, wholesaler_id):
#     wholesaler = Wholesaler.objects.get(id=wholesaler_id)
#     orders = Order.objects.filter(wholesaler_name=wholesaler).order_by('-id')
#     context = {
#         'wholesaler': wholesaler,
#         'orders': orders,
#     }
#     return render(request, 'Admin/wholesaler_view.html', context)





# def wholesaler_view(request, wholesaler_id):
#     wholesaler = Wholesaler.objects.get(id=wholesaler_id)

#     # Get the start and end time from the request parameters
#     start_date_str = request.GET.get('start_date')
#     end_date_str = request.GET.get('end_date')

#     # Convert the start and end time strings to datetime objects
#     start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
#     end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None

#     # Filter the orders based on wholesaler and date range
#     orders = Order.objects.filter(wholesaler_name=wholesaler)
#     if start_date:
#         orders = orders.filter(date__gte=start_date)
#     if end_date:
#         orders = orders.filter(date__lte=end_date)

#     orders = orders.order_by('-id')

#     total_total = sum(order.total for order in orders)

#     # Retrieve the list of purchases for the wholesaler
#     purchases = Purchase.objects.filter(wholesaler_name=wholesaler).order_by('-id')

        
#     if request.method == 'POST':        
#         # Handle Purchase model changes
#         if 'purchase_id' in request.POST:
#             purchase_id = request.POST.get('purchase_id')
#             try:
#                 purchase = Purchase.objects.get(id=purchase_id)
#                 purchase.paid = True
#                 purchase.save()
#                 messages.success(request, 'Purchase marked as paid.')
#             except Purchase.DoesNotExist:
#                 messages.error(request, 'Purchase not found.')
#             except Exception as e:
#                 messages.error(request, str(e))
        
#         # Handle Order model changes
#         if 'order_id' in request.POST:
#             order_id = request.POST.get('order_id')
#             try:
#                 order = Order.objects.get(id=order_id)
#                 order.paid = True
#                 order.save()
#                 messages.success(request, 'Order marked as paid.')
#             except Order.DoesNotExist:
#                 messages.error(request, 'Order not found.')
#             except Exception as e:
#                 messages.error(request, str(e))

#         return redirect('wholesaler_view', wholesaler_id=wholesaler_id)

#     context = {
#         'wholesaler': wholesaler,
#         'orders': orders,
#         'total_total': total_total,
#         'start_date': start_date_str,
#         'end_date': end_date_str,
#         'purchases': purchases,
#     }
#     return render(request, 'Admin/wholesaler_view.html', context)



# def wholesaler_view(request, wholesaler_id):
#     wholesaler = Wholesaler.objects.get(id=wholesaler_id)

#     # Filter the orders based on wholesaler and date range
#     orders = Order.objects.filter(wholesaler_name=wholesaler).order_by('-id')
   

#     # total_total = sum(order.total for order in orders)

#     # Retrieve the list of purchases for the wholesaler
#     purchases = Purchase.objects.filter(wholesaler_name=wholesaler).order_by('-id')

        
#     if request.method == 'POST':        
#         # Handle Purchase model changes
#         if 'purchase_id' in request.POST:
#             purchase_id = request.POST.get('purchase_id')
#             try:
#                 purchase = Purchase.objects.get(id=purchase_id)
#                 purchase.paid = True
#                 purchase.save()
#                 messages.success(request, 'Purchase marked as paid.')
#             except Purchase.DoesNotExist:
#                 messages.error(request, 'Purchase not found.')
#             except Exception as e:
#                 messages.error(request, str(e))
        
#         # Handle Order model changes
#         if 'order_id' in request.POST:
#             order_id = request.POST.get('order_id')
#             try:
#                 order = Order.objects.get(id=order_id)
#                 order.paid = True
#                 order.save()
#                 messages.success(request, 'Order marked as paid.')
#             except Order.DoesNotExist:
#                 messages.error(request, 'Order not found.')
#             except Exception as e:
#                 messages.error(request, str(e))

#         return redirect('wholesaler_view', wholesaler_id=wholesaler_id)

#     context = {
#         'wholesaler': wholesaler,
#         'orders': orders,
#         # 'total_total': total_total,
#         'purchases': purchases,
#     }
#     return render(request, 'Admin/wholesaler_view.html', context)



# def wholesaler_view(request, wholesaler_id):
#     wholesaler = Wholesaler.objects.get(id=wholesaler_id)

#     # Filter the orders based on wholesaler and date range
#     orders = Order.objects.filter(wholesaler_name=wholesaler).order_by('-id')
   

#     # total_total = sum(order.total for order in orders)

#     # Retrieve the list of purchases for the wholesaler
#     purchases = Purchase.objects.filter(wholesaler_name=wholesaler).order_by('-id')

        
#     if request.method == 'POST':        
#         # Handle Purchase model changes
#         if 'purchase_id' in request.POST:
#             purchase_id = request.POST.get('purchase_id')
#             try:
#                 purchase = Purchase.objects.get(id=purchase_id)
#                 purchase.paid = True  # Mark the purchase as paid
#                 purchase.save()
#                 messages.success(request, 'Purchase marked as paid.')
#             except Purchase.DoesNotExist:
#                 messages.error(request, 'Purchase not found.')
#             except Exception as e:
#                 messages.error(request, str(e))
    
#         if 'not_paid_purchase_id' in request.POST:
#             not_paid_purchase_id = request.POST.get('not_paid_purchase_id')
#             try:
#                 purchase = Purchase.objects.get(id=not_paid_purchase_id)
#                 purchase.paid = False  # Mark the purchase as not paid
#                 purchase.save()
#                 messages.success(request, 'Purchase marked as not paid.')
#             except Purchase.DoesNotExist:
#                 messages.error(request, 'Purchase not found.')
#             except Exception as e:
#                 messages.error(request, str(e))
        

#         # Handle Order model changes
#         if 'order_id' in request.POST:
#             order_id = request.POST.get('order_id')
#             try:
#                 order = Order.objects.get(id=order_id)
#                 order.paid = True
#                 order.save()
#                 messages.success(request, 'Order marked as paid.')
#             except Order.DoesNotExist:
#                 messages.error(request, 'Order not found.')
#             except Exception as e:
#                 messages.error(request, str(e))
        

#         if 'not_paid_order_id' in request.POST:
#             not_paid_order_id = request.POST.get('not_paid_order_id')
#             try:
#                 order = Order.objects.get(id=not_paid_order_id)
#                 order.paid = False  # Mark the order as not paid
#                 order.save()
#                 messages.success(request, 'order marked as not paid.')
#             except Order.DoesNotExist:
#                 messages.error(request, 'order not found.')
#             except Exception as e:
#                 messages.error(request, str(e))

#         return redirect('wholesaler_view', wholesaler_id=wholesaler_id)

#     context = {
#         'wholesaler': wholesaler,
#         'orders': orders,
#         # 'total_total': total_total,
#         'purchases': purchases,
#     }
#     return render(request, 'Admin/wholesaler_view.html', context)



# def wholesaler_view(request, wholesaler_id):
#     wholesaler = Wholesaler.objects.get(id=wholesaler_id)
    
#     # Filter the orders based on wholesaler and date range
#     orders = Order.objects.filter(wholesaler_name=wholesaler).order_by('-id')
    
#     debit_total = sum(order.purchase_total for order in orders if not order.paid)
#     credit_total = sum(order.purchase_total for order in orders if order.paid)
#     if request.method == 'POST':        
        
  

#         # Handle Order model changes
#         if 'order_id' in request.POST:
#             order_id = request.POST.get('order_id')
#             try:
#                 order = Order.objects.get(id=order_id)
#                 order.paid = True
#                 order.save()
#                 messages.success(request, ' marked as paid.')
#             except Order.DoesNotExist:
#                 messages.error(request, 'Order not found.')
#             except Exception as e:
#                 messages.error(request, str(e))
        

#         if 'not_paid_order_id' in request.POST:
#             not_paid_order_id = request.POST.get('not_paid_order_id')
#             try:
#                 order = Order.objects.get(id=not_paid_order_id)
#                 order.paid = False  # Mark the order as not paid
#                 order.save()
#                 messages.success(request, ' marked as not paid.')
#             except Order.DoesNotExist:
#                 messages.error(request, 'order not found.')
#             except Exception as e:
#                 messages.error(request, str(e))

#         return redirect('wholesaler_view', wholesaler_id=wholesaler_id)

#     context = {
#         'wholesaler': wholesaler,
#         'orders': orders,
#         'debit_total' : debit_total,
#         'credit_total' : credit_total
        
#     }
#     return render(request, 'Admin/wholesaler_view.html', context)

def wholesaler_view(request, wholesaler_id):
    wholesaler = Wholesaler.objects.get(id=wholesaler_id)
    
    # Filter the orders based on wholesaler and date range
    orders = Order.objects.filter(wholesaler_name=wholesaler).order_by('-id')
    
  
    debit_total = sum(order.purchase_total or 0 for order in orders)

    credit_total = sum(order.wholesaler_paid or 0 for order in orders)

    balance = debit_total - credit_total

    
    

    context = {
        'wholesaler': wholesaler,
        'orders': orders,
        'debit_total' : debit_total,
        'credit_total' : credit_total,
        'balance' : balance
        
    }
    return render(request, 'Admin/wholesaler_view.html', context)

@login_required
def wholesaler_paid(request, wholesaler_id):
    wholesaler = Wholesaler.objects.get(id=wholesaler_id)
    if request.method == 'POST':
        date = request.POST.get('date')
        wholesaler_paid = request.POST.get('wholesaler_paid')
        try:
            Order.objects.create(
                wholesaler_name=wholesaler,
                date=date,
                wholesaler_paid=Decimal(wholesaler_paid)
            )
            messages.success(request, 'Paid added Succesfully')
            return redirect('wholesaler_view', wholesaler_id=wholesaler.id)
        except IntegrityError:
            messages.error(request,'Paid added Field')
            return redirect('wholesaler_paid', wholesaler_id=wholesaler.id)
        except (TypeError, InvalidOperation):
            messages.error(request, "Invalid Value. Please provide a valid number.")
            return redirect('wholesaler_paid', wholesaler_id=wholesaler.id)
    return render(request, 'Admin/wholesaler_paid.html')

@login_required
def wholesaler_delete(request, wholesaler_id):
    try:
        wholesaler = Wholesaler.objects.get(id=wholesaler_id)
    except wholesaler.DoesNotExist:
        messages.error(request, f"wholesaler with ID {wholesaler_id} does not exist.")
        return redirect('wholesaler_list')
    
    try:
        wholesaler.delete()
        messages.success(request, f"wholesaler '{wholesaler.name}' Delete successfully.")
        return redirect('wholesaler_list')
    except IntegrityError:
            messages.error(request, f"The username '{wholesaler.name}' Delete Failed.")


# ---------------------------------------------- Order -------------------------------------------------------- #

# @login_required
# def order_list(request):
#     order = Order.objects.all().order_by('-id')
#     context = {
#         'order' : order,
#     }
#     return render(request, 'Admin/order_list.html', context)

@login_required
def order_list(request):
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    

    # Convert the start and end time strings to datetime objects
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None

    # Filter the orders based on date range and customer
    order = Order.objects.exclude(customer_name=None).order_by('-id')
    if start_date:
        order = order.filter(date__gte=start_date)
    if end_date:
        order = order.filter(date__lte=end_date)
    context = {
        'order' : order,
        'start_date': start_date_str,
        'end_date': end_date_str,
    }
    return render(request, 'Admin/order_list.html', context)


@login_required
def order_add(request):
    customers = Customer.objects.all()
    wholesalers = Wholesaler.objects.all()
    collectors = User.objects.filter(is_collector=True)
    delivery_boys = User.objects.filter(is_delivery_boy=True)
    if request.method == 'POST':
        date = request.POST.get('date') 
        customer_id = request.POST.get('customer_name') 

        # money_type = request.POST.get('money_type')

        purchase_mrp = request.POST.get('purchase_mrp') 
        purchase_quantity = request.POST.get('purchase_quantity') 
        purchase_total = request.POST.get('purchase_total')

        mrp = request.POST.get('mrp') 
        quantity = request.POST.get('quantity') 
        total = request.POST.get('total') 

        collector_amount = request.POST.get('collector_amount') 
        collector_id = request.POST.get('collector_name') 
        name = request.POST.get('name') 
        mobile = request.POST.get('mobile') 
        address = request.POST.get('address')

        delivery_type = request.POST.get('delivery_type')

        wholesaler_id = request.POST.get('wholesaler_name') 

        delivery_boy_id = request.POST.get('delivery_boy_name') 
        delivery_boy_amount = request.POST.get('delivery_boy_amount')

        home_name = request.POST.get('home_name') 
        home_mobile = request.POST.get('home_mobile') 
        home_address = request.POST.get('home_address')

        person_name = request.POST.get('person_name') 
        bank_name = request.POST.get('bank_name') 
        account = request.POST.get('account') 
        ifse = request.POST.get('ifse') 
        

        customer = Customer.objects.get(id=customer_id)

        if wholesaler_id:
            wholesaler = Wholesaler.objects.get(id=wholesaler_id)
        else:
            wholesaler = None

        collector = User.objects.filter(id=collector_id).first()
        if delivery_boy_id:
            delivery_boy = User.objects.filter(id=delivery_boy_id).first()
        else:
            delivery_boy = None
        try:
            order = Order.objects.create(
                date=date,
                customer_name=customer, 

                # money_type=money_type, 

                purchase_mrp=Decimal(purchase_mrp),
                purchase_quantity=purchase_quantity,
                purchase_total=Decimal(purchase_total),

                mrp=Decimal(mrp),
                quantity=quantity,
                total=Decimal(total),

                collector_amount=collector_amount,
                collector_name=collector,  
                name=name,
                mobile=mobile,
                address=address,

                delivery_type=delivery_type,

                wholesaler_name=wholesaler,

                delivery_boy_name=delivery_boy,
                delivery_boy_amount=delivery_boy_amount,

                home_name=home_name,
                home_mobile=home_mobile,
                home_address=home_address,

                person_name=person_name,
                bank_name=bank_name,
                account=account,
                ifse=ifse,
                delivery_status='PENDING',
                comment = '',
                
                money_collected = 0,
                money_pending = None
            )
            messages.success(request, f"Order added successfully.")
            return redirect('order_list')
        except IntegrityError:
            messages.error(request, f"Order add Failed.")
            return redirect('order_add')
        except (TypeError, InvalidOperation):
            messages.error(request, "Invalid Value. Please provide a valid number.")
            return redirect('order_add')
        
    context = {
        'customers': customers,
        'wholesalers': wholesalers,
        'collectors': collectors,
        'delivery_boys': delivery_boys,
    }
    return render(request, 'Admin/order_add.html', context)
        


# for automatic field add

def get_customer_details(request):
    customer_id = request.GET.get('customer_id')
    show_details = request.GET.get('show_details', False)  # Get the checkbox status
    try:
        customer = Customer.objects.get(pk=customer_id)
        customer_data = {
            'name': customer.name if show_details else '',
            'mobile': customer.mobile if show_details else '',
            'address': customer.address if show_details else '',
        }
        return JsonResponse(customer_data)
    except Customer.DoesNotExist:
        return JsonResponse({'error': 'Customer not found'}, status=404)
    


@login_required
def order_edit(request, order_id):
    try:
        order = Order.objects.get(id=order_id)
    except Order.DoesNotExist:
        messages.error(request, f"Order with ID {order_id} does not exist.")
        return redirect('order_list')

    customers = Customer.objects.all()
    wholesalers = Wholesaler.objects.all()
    collectors = User.objects.filter(is_collector=True)
    delivery_boys = User.objects.filter(is_delivery_boy=True)

    if request.method == 'POST':
        order.date = request.POST.get('date')

        customer_id = request.POST.get('customer_name') 
        customer_instance = Customer.objects.get(name=customer_id)
        order.customer_name = customer_instance

        order.purchase_mrp = request.POST.get('purchase_mrp') 
        order.purchase_quantity = request.POST.get('purchase_quantity') 
        order.purchase_total = request.POST.get('purchase_total') 

        order.mrp = request.POST.get('mrp') 
        order.quantity = request.POST.get('quantity') 
        order.total = request.POST.get('total') 

        order.collector_amount = request.POST.get('collector_amount')

        collector_id = request.POST.get('collector_name')
        collector_instance = User.objects.get(id=collector_id)
        order.collector_name = collector_instance

        order.name = request.POST.get('name') 
        order.mobile = request.POST.get('mobile') 
        order.address = request.POST.get('address')


        delivery_boy_id = request.POST.get('delivery_boy_name')
        if delivery_boy_id:
            delivery_boy_instance = User.objects.get(id=delivery_boy_id)
            order.delivery_boy_name = delivery_boy_instance
        else:
            order.delivery_boy_name = None

        wholesaler_id = request.POST.get('wholesaler_name')
        if wholesaler_id:
            wholesaler_instance = Wholesaler.objects.get(name=wholesaler_id)
            order.wholesaler_name = wholesaler_instance
        else:
            order.wholesaler_name = None  # Corrected line
    
        order.delivery_boy_amount = request.POST.get('delivery_boy_amount')

        order.home_name = request.POST.get('home_name') 
        order.home_mobile = request.POST.get('home_mobile') 
        order.home_address = request.POST.get('home_address')

        order.person_name = request.POST.get('person_name') 
        order.bank_name = request.POST.get('bank_name') 
        order.account = request.POST.get('account') 
        order.ifse = request.POST.get('ifse') 

        # for delivery boys
        order.delivery_status = 'PENDING'
        order.comment = ''
        
        order.money_collected = 0
        order.money_pending = None
        try:
            order.save()
            messages.success(request, f"Order updated successfully.")
            return redirect('order_list')
        except IntegrityError:
            messages.error(request, f"Order add Failed.")
            return redirect('order_edit', order_id=order_id)

    context = {
        'order': order,
        'customers': customers,
        'wholesalers': wholesalers,
        'collectors': collectors,
        'delivery_boys': delivery_boys,
    }
    return render(request, 'Admin/order_edit.html', context)

@login_required
def order_view(request, order_id):
    order = Order.objects.get(id=order_id)
    context = {
        'order' : order,
    }
    return render(request, 'Admin/order_view.html', context)    

@login_required
def order_delete(request, order_id):
    try:
        order = Order.objects.get(id=order_id)
    except Order.DoesNotExist:
        messages.error(request, f"Order with ID {order_id} does not exist.")
        return redirect('order_list')

    try:
        order.delete()
        messages.success(request, f" Delete successfully.")
        return redirect('order_list')
    except IntegrityError:
            messages.error(request, f" Delete Failed.")
            return redirect('order_list')
    

# ---------------------------------------------- Report -------------------------------------------------------- #


# @login_required
# def report(request):
#     # Get the start and end time from the request parameters
#     start_date_str = request.GET.get('start_date')
#     end_date_str = request.GET.get('end_date')

#     # Convert the start and end time strings to datetime objects
#     start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
#     end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None

#     # Filter the orders based on date range
#     orders = Order.objects.all().order_by('-id')
#     if start_date:
#         orders = orders.filter(date__gte=start_date)
#     if end_date:
#         orders = orders.filter(date__lte=end_date)

#     # Count the total orders within the selected date range
#     order_count = orders.count()

#     profit_count = sum(order.profit for order in orders )
    
#     total_money_collected = sum(order.money_collected if order.money_collected is not None else 0 for order in orders)
#     total_money_pending = sum(order.money_pending if order.money_pending is not None else order.total for order in orders)

#     context = {
#         'orders': orders,
#         'start_date': start_date_str,
#         'end_date': end_date_str,
#         'order_count': order_count,
#         'profit_count' : profit_count,
#         'total_money_collected' : total_money_collected,
#         'total_money_pending' : total_money_pending
        
        
#     }
#     return render(request, 'Admin/report.html', context)



# @login_required
# def report(request):
#     customers = Customer.objects.all()
#     # Get the start and end time from the request parameters
#     start_date_str = request.GET.get('start_date')
#     end_date_str = request.GET.get('end_date')
    
#     # Get the customer ID from the request parameters
#     customer_id = request.GET.get('customer_id')  # Assuming the parameter name is 'customer_id'

#     # Convert the start and end time strings to datetime objects
#     start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
#     end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None

#     # Filter the orders based on date range and customer
#     orders = Order.objects.all().order_by('-id')
#     if start_date:
#         orders = orders.filter(date__gte=start_date)
#     if end_date:
#         orders = orders.filter(date__lte=end_date)
#     if customer_id:
#         orders = orders.filter(customer_name__id=customer_id)

#     # Count the total orders within the selected date range and customer
#     order_count = orders.count()

#     profit_count = sum(order.profit for order in orders)
#     total_money_collected = sum(order.money_collected if order.money_collected is not None else 0 for order in orders)
#     total_money_pending = sum(order.money_pending if order.money_pending is not None else order.total for order in orders)

#     context = {
#         'orders': orders,
#         'start_date': start_date_str,
#         'end_date': end_date_str,
#         'order_count': order_count,
#         'profit_count': profit_count,
#         'total_money_collected': total_money_collected,
#         'total_money_pending': total_money_pending,
#         'customers': customers,
#     }
#     return render(request, 'Admin/report.html', context)



# @login_required
# def report(request):
#     customers = Customer.objects.all()
#     # Get the start and end time from the request parameters
#     start_date_str = request.GET.get('start_date')
#     end_date_str = request.GET.get('end_date')
    
#     # Get the customer ID from the request parameters
#     customer_id = request.GET.get('customer_id')  # Assuming the parameter name is 'customer_id'

#     # Convert the start and end time strings to datetime objects
#     start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
#     end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None

#     # Filter the orders based on date range and customer
#     orders = Order.objects.all().order_by('-id')
#     if start_date:
#         orders = orders.filter(date__gte=start_date)
#     if end_date:
#         orders = orders.filter(date__lte=end_date)
#     if customer_id:
#         orders = orders.filter(customer_name__id=customer_id)

#     # Count the total orders within the selected date range and customer
#     order_count = orders.count()

#     profit_count = sum(order.profit for order in orders)
#     total_money_collected = sum(order.money_collected if order.money_collected is not None else 0 for order in orders)
#     total_money_pending = sum(order.money_pending if order.money_pending is not None else order.total for order in orders)

#     if 'export' in request.GET:
#                 response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#                 response['Content-Disposition'] = 'attachment; filename=orders.xlsx'

#                 workbook = Workbook()
#                 worksheet = workbook.active

#                 # Write headers
#                 headers = ['Customer', 'Date', 'Purchase Total', 'Sales Total', 'Profit', 'Collector Agent', 'Delivery Agent', 'Collected', 'Pending',]
#                 for col_num, header in enumerate(headers, 1):
#                     worksheet.cell(row=1, column=col_num, value=header)

#                 for col_num in range(1, len(headers) + 1):
#                     column_letter = get_column_letter(col_num)
#                     worksheet.column_dimensions[column_letter].width = 15

#                 # Write data
#                 for row_num, order in enumerate(orders, 2):
#                     worksheet.cell(row=row_num, column=1, value=order.customer_name.name)  # Assuming name is the field representing customer name
#                     worksheet.cell(row=row_num, column=2, value=order.date.strftime('%Y-%m-%d'))  # Convert date to formatted string
#                     worksheet.cell(row=row_num, column=3, value=order.purchase_total)
#                     worksheet.cell(row=row_num, column=4, value=order.total)
#                     worksheet.cell(row=row_num, column=5, value=order.profit)

#                     collector_full_name = f"{order.collector_name.first_name} {order.collector_name.last_name}"
#                     worksheet.cell(row=row_num, column=6, value=collector_full_name)
                    
#                     delivery_boy_full_name = f"{order.delivery_boy_name.first_name} {order.delivery_boy_name.last_name}"
#                     worksheet.cell(row=row_num, column=7, value=delivery_boy_full_name)

#                     worksheet.cell(row=row_num, column=8, value=order.money_collected)

#                     if order.money_pending == None:
#                         cell_value = order.total
#                     else:
#                         cell_value = order.money_pending
#                     worksheet.cell(row=row_num, column=9, value=cell_value)

#                 workbook.save(response)
#                 return response

#     context = {
#         'orders': orders,
#         'start_date': start_date_str,
#         'end_date': end_date_str,
#         'order_count': order_count,
#         'profit_count': profit_count,
#         'total_money_collected': total_money_collected,
#         'total_money_pending': total_money_pending,
#         'customers': customers,
#     }
#     return render(request, 'Admin/report.html', context)



def report(request):
    customers = Customer.objects.all()
    # Get the start and end time from the request parameters
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    # Get the customer ID from the request parameters
    customer_id = request.GET.get('customer_id')

    # Convert the start and end time strings to datetime objects
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None

    # dont get customer is none
    orders = Order.objects.exclude(customer_name=None).order_by('-id')
    if start_date:
        orders = orders.filter(date__gte=start_date)
    if end_date:
        orders = orders.filter(date__lte=end_date)
    if customer_id:
        orders = orders.filter(customer_name__id=customer_id)

    # orders = orders.exclude(wholesaler_name__isnull=False)

    


    # Count the total orders within the selected date range and customer
    order_count = orders.count()
    profit_count = sum(order.profit for order in orders)
    total_money_collected = sum(order.money_collected if order.money_collected is not None else 0 for order in orders)
    total_money_pending = sum(order.money_pending if order.money_pending is not None else order.total for order in orders)
    total_sale = sum(order.total for order in orders)




    if 'export' in request.GET:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=orders.xlsx'

        workbook = Workbook()
        worksheet = workbook.active

        # Write headers
        headers = ['Customer', 'Date', 'Purchase Total', 'Sales Total', 'Profit', 'Collector Agent', 'Delivery Agent', 'Collected', 'Pending']
        for col_num, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_num, value=header)

        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = 15

        # Write data
        for row_num, order in enumerate(orders, 2):
            worksheet.cell(row=row_num, column=1, value=order.customer_name.name)
            worksheet.cell(row=row_num, column=2, value=order.date.strftime('%Y-%m-%d'))
            worksheet.cell(row=row_num, column=3, value=order.purchase_total)
            worksheet.cell(row=row_num, column=4, value=order.total)
            worksheet.cell(row=row_num, column=5, value=order.profit)

            collector_full_name = f"{order.collector_name.first_name} {order.collector_name.last_name}"
            worksheet.cell(row=row_num, column=6, value=collector_full_name)
                    
            delivery_boy_full_name = f"{order.delivery_boy_name.first_name} {order.delivery_boy_name.last_name}"
            worksheet.cell(row=row_num, column=7, value=delivery_boy_full_name)

            worksheet.cell(row=row_num, column=8, value=order.money_collected)

            if order.money_pending is None:
                cell_value = order.total
            else:
                cell_value = order.money_pending
            worksheet.cell(row=row_num, column=9, value=cell_value)

        workbook.save(response)
        return response

    context = {
        'orders': orders,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'order_count': order_count,
        'profit_count': profit_count,
        'total_sale' : total_sale,
        'total_money_collected': total_money_collected,
        'total_money_pending': total_money_pending,
        'customers': customers,
        'selected_customer_id': int(customer_id) if customer_id else None,
    }
    return render(request, 'Admin/report.html', context)




# ---------------------------------------------- Delivery Boy Salery -------------------------------------------------------- #

@login_required
def salary_delivery_boy_list(request):
    delivery_boy = User.objects.filter(is_delivery_boy=True)
    context = {
        'delivery_boy' : delivery_boy,
    }
    return render(request, 'Admin/salary_delivery_boy_list.html', context)


@login_required
def salary_delivery_boy_salary_list(request, delivery_boy_id):
    delivery_boy = User.objects.get(id=delivery_boy_id)

    if request.method == 'POST':
        date = request.POST.get('date')
        salary_amount = request.POST.get('salary_amount')

        try:
            salary_amount = Decimal(salary_amount)
        except (TypeError, InvalidOperation):
            messages.error(request, 'Invalid salary amount. Please provide a valid number.')
        else:
            
            if salary_amount <= 0:
                messages.error(request, 'Salary amount should be a positive number.')
            else:
                DeliveryBoySalary.objects.create(
                    delivery_boy_name=delivery_boy,
                    date=date,
                    salary_amount=salary_amount
                )

    
    salary = DeliveryBoySalary.objects.filter(delivery_boy_name=delivery_boy).order_by('-id')

    context = {
        'salary': salary,
        'delivery_boy': delivery_boy,
    }
    return render(request, 'Admin/salary_delivery_boy_salary_list.html', context)


@login_required
def salary_delivery_boy_delete(request, salary_id):
    try:
        salary = DeliveryBoySalary.objects.get(id=salary_id)
    except Order.DoesNotExist:
        messages.error(request, f"Order with ID {salary_id} does not exist.")
        return redirect('salary_delivery_boy_salary_list', delivery_boy_id=salary.delivery_boy_name.id)

    try:
        salary.delete()
        messages.success(request, f"Salry Delete successfully.")
        return redirect('salary_delivery_boy_salary_list', delivery_boy_id=salary.delivery_boy_name.id)
    except IntegrityError:
            messages.error(request, f"Salary Delete Failed.")
            return redirect('salary_delivery_boy_salary_list', delivery_boy_id=salary.delivery_boy_name.id)
    

# ---------------------------------------------- Purchase -------------------------------------------------------- #



def purchase_add(request, wholesaler_id):
    try:
        wholesaler = Wholesaler.objects.get(id=wholesaler_id)
    except Wholesaler.DoesNotExist:
        messages.error(request, f"Wholesaler with ID {wholesaler_id} does not exist.")
        return redirect('wholesaler_view', wholesaler_id=wholesaler_id)

    delivery_boys = User.objects.filter(is_delivery_boy=True)

    if request.method == 'POST':
        # Handle form submission
        delivery_boy_id = request.POST.get('delivery_boy_name')
        date = request.POST.get('date')
        purchase_mrp = request.POST.get('mrp')
        purchase_quantity = request.POST.get('quantity')
        purchase_total = request.POST.get('total')

        try:
            delivery_boy = User.objects.get(id=delivery_boy_id)
        except User.DoesNotExist:
            messages.error(request, f"Please select a Delivery Agent")
            return redirect('purchase_add', wholesaler_id=wholesaler_id)

        # Create and save the Purchase object
        try:
            purchase = Order(
                wholesaler_name=wholesaler,
                delivery_boy_name=delivery_boy,
                purchase_mrp=Decimal(purchase_mrp),
                purchase_quantity=purchase_quantity,
                purchase_total=Decimal(purchase_total),
                date=date,
                wholesaler_show = 'SHOW',

                
            )
            purchase.save()
            messages.success(request, f"Purchase Add successfully.")
            return redirect('wholesaler_view', wholesaler_id=wholesaler_id)
        except IntegrityError:
            messages.error(request, f" Failed.")
            return redirect('purchase_add')
        except (TypeError, InvalidOperation):
            messages.error(request, "Invalid Value. Please provide a valid number.")
            return redirect('purchase_add', wholesaler_id=wholesaler_id)

       

    context = {
        'delivery_boys': delivery_boys,
        'wholesaler': wholesaler,  # Pass the 'wholesaler' object to the template context
    }
    return render(request, 'Admin/purchase_add.html', context)






